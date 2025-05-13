import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple  # Import the correct function
from datetime import datetime, timedelta

class planilhaContagem:
    def __init__(self, codigo="", ponto=""):
        self.filename = f"{codigo}_{ponto}.xlsx"
        self.wb = Workbook()
        self.vehicle_data = []  # Store data for summary
        self.entrada = self.abaEntrada(self.wb)
        self.resumo = self.abaResumo(self.wb)
        self.relatorio = self.abaRelatorio(self.wb, self)  # Pass parent
        self.hr = self.abaHr(self.wb, self)  # Pass parent

    def column_to_number(self, col_str):
        """Convert Excel column letters (e.g., 'A', 'AD') to numerical index (1-based)."""
        number = 0
        for char in col_str.upper():
            number = number * 26 + (ord(char) - ord('A') + 1)
        return number

    class abaEntrada:
        def __init__(self, wb):
            self.wb = wb
            self.sheet1 = self.wb.active
            self.sheet1.title = "Entrada"
            self.header_font = Font(bold=True)
            self.title_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.center_align = Alignment(horizontal='center', vertical='center')

        def add_data(self, data=None):
            headers = [
                ('B1', "Ponto"), ('B2', "Data inicial"), ('B3', "Num_Movimentos"),
                ('B4', "Localização"), ('B5', "Duração_dias"), ('B6', "Duração_horas"),
                ('B7', "Hora_início"), ('B8', "Hora_fim"), ('E1', "Movimento")
            ]
            for cell_pos, value in headers:
                cell = self.sheet1[cell_pos]
                cell.value = value
                cell.font = self.header_font
                cell.fill = self.title_fill
                cell.border = self.border
                cell.alignment = self.center_align

            if data:
                data_cells = [
                    ('C1', data.get("Ponto", "")),
                    ('C2', data.get("Data", "")),
                    ('C3', data.get("Num_Movimentos")),
                    ('C4', data.get("Localização")),
                    ('C5', data.get("Duração em dias")),
                    ('C6', data.get("Duração em horas")),
                    ('C7', data.get("Periodo_Inicio", "")),
                    ('C8', data.get("Periodo_Fim", ""))
                ]
                for cell_pos, value in data_cells:
                    cell = self.sheet1[cell_pos]
                    cell.value = value
                    cell.border = self.border
                    cell.alignment = self.center_align

                movimentos = data.get("Movimentos", [])
                for i, movimento in enumerate(movimentos, start=2):
                    cell = self.sheet1[f'E{i}']
                    cell.value = movimento
                    cell.border = self.border
                    cell.alignment = self.center_align

    class abaResumo:
        def __init__(self, wb):
            self.wb = wb
            self.sheet = self.wb.create_sheet(title="Resumo")
            self.header_font = Font(bold=True)
            self.header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.center_align = Alignment(horizontal='center', vertical='center')

        def add_data(self, vehicle_data, days):
            entrada_sheet = self.wb['Entrada']
            ponto = entrada_sheet['C1'].value or "Unknown"
            num_movimentos = entrada_sheet['C3'].value or 0
            duracao_horas = entrada_sheet['C6'].value or 0

            try:
                num_movimentos = int(num_movimentos)
            except (ValueError, TypeError):
                num_movimentos = 0
                print(f"Warning: Invalid 'Num_Movimentos' value '{entrada_sheet['C3'].value}', defaulting to 0.")

            movimentos = []
            if num_movimentos > 0:
                movimentos = [
                    entrada_sheet[f'E{i}'].value
                    for i in range(2, 2 + num_movimentos)
                    if entrada_sheet[f'E{i}'].value
                ]

            self.sheet['B2'] = "Resumo por Categoria"
            self.sheet['B2'].font = self.header_font
            self.sheet['B2'].fill = self.header_fill
            self.sheet['B2'].border = self.border
            headers = ['Movimento', 'Quantidade', 'Duração da Contagem (horas)']
            for col, header in enumerate(headers, 2):
                cell = self.sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            if movimentos:
                for row, movimento in enumerate(movimentos, 4):
                    cell = self.sheet.cell(row=row, column=2)
                    cell.value = movimento
                    cell.border = self.border
                    cell.alignment = self.center_align
                    cell = self.sheet.cell(row=row, column=3)
                    cell.value = 1
                    cell.border = self.border
                    cell.alignment = self.center_align
                    cell = self.sheet.cell(row=row, column=4)
                    cell.value = duracao_horas
                    cell.border = self.border
                    cell.alignment = self.center_align
            else:
                cell = self.sheet.cell(row=4, column=2)
                cell.value = "Nenhum movimento registrado"
                cell.border = self.border
                cell.alignment = self.center_align

            self.sheet['B10'] = "Horário Pico Manhã (7h–8h)"
            self.sheet['B10'].font = self.header_font
            self.sheet['B10'].fill = self.header_fill
            self.sheet['B10'].border = self.border
            vehicle_types = ['Leves', 'VUC', 'Caminhões', 'Carretas', 'Ônibus', 'Pesados', 'Motos', 'Total', 'Total s/VUC']
            headers = ['Movimento'] + vehicle_types
            for col, header in enumerate(headers, 2):
                cell = self.sheet.cell(row=11, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            try:
                hr_sheet = self.wb['Hr']
                if movimentos:
                    for row, movimento in enumerate(movimentos, 12):
                        cell = self.sheet.cell(row=row, column=2)
                        cell.value = movimento
                        cell.border = self.border
                        cell.alignment = self.center_align
                        hr_row = 5 + 7
                        col_mapping = {
                            'Leves': 'D', 'VUC': 'H', 'Caminhões': 'W', 'Carretas': 'Y',
                            'Ônibus': 'AA', 'Motos': 'U'
                        }
                        for col_idx, vt in enumerate(vehicle_types, 3):
                            cell = self.sheet.cell(row=row, column=col_idx)
                            col_letter = col_mapping.get(vt)
                            if col_letter:
                                cell.value = f"='Hr'!{col_letter}{hr_row}"
                            elif vt == 'Pesados':
                                cell.value = f"=SUM(E{row},F{row},G{row})"
                            elif vt == 'Total':
                                cell.value = f"=SUM(C{row},D{row},I{row})"
                            elif vt == 'Total s/VUC':
                                cell.value = f"=SUM(C{row},I{row})"
                            cell.border = self.border
                            cell.alignment = self.center_align
                else:
                    cell = self.sheet.cell(row=12, column=2)
                    cell.value = "Nenhum movimento registrado"
                    cell.border = self.border
                    cell.alignment = self.center_align
            except KeyError:
                cell = self.sheet.cell(row=12, column=2)
                cell.value = "Folha 'Hr' não encontrada"
                cell.border = self.border
                cell.alignment = self.center_align
                print("Warning: 'Hr' sheet not found in workbook.")

            if movimentos:
                chart = BarChart()
                chart.title = "Horário Pico Manhã (7h–8h) Distribuição de veículos"
                chart.x_axis.title = "Movimento"
                chart.y_axis.title = "Number of Vehicles"
                data = Reference(self.sheet, min_col=3, max_col=11, min_row=11, max_row=11+len(movimentos))
                cats = Reference(self.sheet, min_col=2, min_row=12, max_row=12+len(movimentos)-1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                self.sheet.add_chart(chart, "M5")

            self.sheet['B20'] = "Horário Pico Tarde (17h–18h)"
            self.sheet['B20'].font = self.header_font
            self.sheet['B20'].fill = self.header_fill
            self.sheet['B20'].border = self.border
            for col, header in enumerate(headers, 2):
                cell = self.sheet.cell(row=21, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            try:
                hr_sheet = self.wb['Hr']
                if movimentos:
                    for row, movimento in enumerate(movimentos, 22):
                        cell = self.sheet.cell(row=row, column=2)
                        cell.value = movimento
                        cell.border = self.border
                        cell.alignment = self.center_align
                        hr_row = 5 + 17
                        for col_idx, vt in enumerate(vehicle_types, 3):
                            cell = self.sheet.cell(row=row, column=col_idx)
                            col_letter = col_mapping.get(vt)
                            if col_letter:
                                cell.value = f"='Hr'!{col_letter}{hr_row}"
                            elif vt == 'Pesados':
                                cell.value = f"=SUM(E{row},F{row},G{row})"
                            elif vt == 'Total':
                                cell.value = f"=SUM(C{row},D{row},I{row})"
                            elif vt == 'Total s/VUC':
                                cell.value = f"=SUM(C{row},I{row})"
                            cell.border = self.border
                            cell.alignment = self.center_align
                else:
                    cell = self.sheet.cell(row=22, column=2)
                    cell.value = "Nenhum movimento registrado"
                    cell.border = self.border
                    cell.alignment = self.center_align
            except KeyError:
                cell = self.sheet.cell(row=22, column=2)
                cell.value = "Folha 'Hr' não encontrada"
                cell.border = self.border
                cell.alignment = self.center_align
                print("Warning: 'Hr' sheet not found in workbook.")

            if movimentos:
                chart = BarChart()
                chart.title = "Horário Pico Tarde (17h–18h) Distribuição de veículos"
                chart.x_axis.title = "Movimento"
                chart.y_axis.title = "Number of Vehicles"
                data = Reference(self.sheet, min_col=3, max_col=11, min_row=21, max_row=21+len(movimentos))
                cats = Reference(self.sheet, min_col=2, min_row=22, max_row=22+len(movimentos)-1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                self.sheet.add_chart(chart, "M20")

            start_row = 30 + len(movimentos)
            self.sheet[f'B{start_row}'] = "Fluxo Total do Dia"
            self.sheet[f'B{start_row}'].font = self.header_font
            self.sheet[f'B{start_row}'].fill = self.header_fill
            self.sheet[f'B{start_row}'].border = self.border
            vehicle_types_daily = ['Leves', 'VUC', 'Caminhões', 'Carretas', 'Ônibus', 'Pesados', 'Motos', 'Total']
            headers_daily = ['Movimento'] + vehicle_types_daily
            for col, header in enumerate(headers_daily, 2):
                cell = self.sheet.cell(row=start_row+1, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            try:
                hr_sheet = self.wb['Hr']
                if movimentos:
                    for idx, movimento in enumerate(movimentos, 0):
                        row = start_row + 2 + idx
                        cell = self.sheet.cell(row=row, column=2)
                        cell.value = movimento
                        cell.border = self.border
                        cell.alignment = self.center_align
                        hr_footer_row = 28 + (idx * 6)
                        col_mapping_daily = {
                            'Leves': 'D', 'VUC': 'H', 'Caminhões': 'W', 'Carretas': 'Y',
                            'Ônibus': 'AA', 'Pesados': 'AC', 'Motos': 'U', 'Total': 'AD'
                        }
                        for col_idx, vt in enumerate(vehicle_types_daily, 3):
                            cell = self.sheet.cell(row=row, column=col_idx)
                            col_letter = col_mapping_daily.get(vt)
                            if col_letter:
                                cell.value = f"='Hr'!{col_letter}{hr_footer_row}"
                            cell.border = self.border
                            cell.alignment = self.center_align
                else:
                    cell = self.sheet.cell(row=start_row+2, column=2)
                    cell.value = "Nenhum movimento registrado"
                    cell.border = self.border
                    cell.alignment = self.center_align
            except KeyError:
                cell = self.sheet.cell(row=start_row+2, column=2)
                cell.value = "Folha 'Hr' não encontrada"
                cell.border = self.border
                cell.alignment = self.center_align
                print("Warning: 'Hr' sheet not found in workbook.")

            if movimentos:
                chart = BarChart()
                chart.title = "Fluxo Total do Dia por Movimento"
                chart.x_axis.title = "Movimento"
                chart.y_axis.title = "Number of Vehicles"
                data = Reference(self.sheet, min_col=3, max_col=10, min_row=start_row+1, max_row=start_row+1+len(movimentos))
                cats = Reference(self.sheet, min_col=2, min_row=start_row+2, max_row=start_row+2+len(movimentos)-1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                self.sheet.add_chart(chart, f"M{start_row}")

            for col in self.sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min((max_length + 2), 50)
                self.sheet.column_dimensions[column].width = adjusted_width

    class abaRelatorio:
        def __init__(self, wb, parent):
            self.wb = wb
            self.parent = parent
            self.sheet2 = self.wb.create_sheet(title="Relatório")
            self.header_font = Font(bold=True, size=11)
            self.header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            self.pesados_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.separator_border = Border(
                left=Side(style='medium'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.center_align = Alignment(horizontal='center', vertical='center')

        def create_movement_table(self, start_row, data, movement, movement_index):
            ponto = data.get("Ponto", "")
            movimento_concatenado = f"{ponto}{movement}" if ponto and movement else movement

            cell = self.sheet2[f'B{start_row}']
            cell.value = "Data:"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align

            cell = self.sheet2[f'C{start_row}']
            cell.value = data.get("Data", "")
            cell.border = self.border
            cell.alignment = self.center_align

            cell = self.sheet2[f'B{start_row + 1}']
            cell.value = "Movimento:"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align

            cell = self.sheet2[f'C{start_row + 1}']
            cell.value = movimento_concatenado
            cell.border = self.border
            cell.alignment = self.center_align

            header_row = start_row + 2
            subcat_row = start_row + 3
            table_columns = [
                (f'B{header_row}:C{header_row}', "Horas", None),
                (f'D{header_row}:D{subcat_row}', "Leves", None),
                (f'E{header_row}:G{header_row}', "Carretinha", None),
                (f'H{header_row}:H{subcat_row}', "VUC", None),
                (f'I{header_row}:K{header_row}', "Caminhões", None),
                (f'L{header_row}:S{header_row}', "Carreta", None),
                (f'T{header_row}:U{header_row}', "Ônibus", None),
                (f'V{header_row}:V{subcat_row}', "Motos", None),
                (f'W{header_row}:AD{header_row}', "Pesados", None),
                (f'AE{header_row}:AE{subcat_row}', "Veículos Totais", None)
            ]

            # Step 1: Apply borders, fill, and values to individual cells before merging
            for merge_range, header_value, _ in table_columns:
                start_cell, end_cell = merge_range.split(':')
                start_row_num, start_col = coordinate_to_tuple(start_cell)
                end_row_num, end_col = coordinate_to_tuple(end_cell)
                for row in range(start_row_num, end_row_num + 1):
                    for col in range(start_col, end_col + 1):
                        col_letter = get_column_letter(col)
                        cell = self.sheet2[f'{col_letter}{row}']
                        # Apply fill and border based on whether this cell is in the "Pesados" section
                        if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'] and row == header_row:
                            cell.fill = self.pesados_fill
                            cell.border = self.border
                        elif col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'] and row == subcat_row:
                            cell.fill = self.pesados_fill
                            cell.border = self.border
                        elif col_letter == 'V' and row in [header_row, subcat_row]:
                            cell.fill = self.pesados_fill
                            cell.border = self.separator_border
                        else:
                            if row == header_row:
                                cell.fill = self.header_fill
                            cell.border = self.border
                        # Set the header value only for the top-left cell
                        if row == start_row_num and col == start_col and header_value:
                            cell.value = header_value
                            cell.font = self.header_font
                            cell.alignment = self.center_align

            # Step 2: Merge the cells
            for merge_range, _, _ in table_columns:
                self.sheet2.merge_cells(merge_range)

            # Step 3: Reapply border and fill to the top-left cell of each merged area to ensure outer borders are visible
            for merge_range, _, _ in table_columns:
                start_cell, end_cell = merge_range.split(':')
                start_row_num, start_col = coordinate_to_tuple(start_cell)
                end_row_num, end_col = coordinate_to_tuple(end_cell)
                # Reapply to the top-left cell of the merged area
                cell = self.sheet2[start_cell]
                col_letter = start_cell[0]
                if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                    cell.fill = self.pesados_fill
                    cell.border = self.border
                elif col_letter == 'V':
                    cell.fill = self.pesados_fill
                    cell.border = self.separator_border
                else:
                    cell.fill = self.header_fill
                    cell.border = self.border

                # Ensure the outer borders of the merged area are visible by reapplying borders to the edges
                for row in range(start_row_num, end_row_num + 1):
                    # Leftmost column of the merged area
                    left_col_letter = get_column_letter(start_col)
                    cell_left = self.sheet2[f'{left_col_letter}{row}']
                    if left_col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_left.border = self.border
                    elif left_col_letter == 'V':
                        cell_left.border = self.separator_border
                    else:
                        cell_left.border = self.border

                    # Rightmost column of the merged area
                    right_col_letter = get_column_letter(end_col)
                    cell_right = self.sheet2[f'{right_col_letter}{row}']
                    if right_col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_right.border = self.border
                    else:
                        cell_right.border = self.border

                # Top and bottom rows of the merged area
                for col in range(start_col, end_col + 1):
                    col_letter = get_column_letter(col)
                    # Top row
                    cell_top = self.sheet2[f'{col_letter}{start_row_num}']
                    if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_top.border = self.border
                    elif col_letter == 'V':
                        cell_top.border = self.separator_border
                    else:
                        cell_top.border = self.border

                    # Bottom row
                    cell_bottom = self.sheet2[f'{col_letter}{end_row_num}']
                    if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_bottom.border = self.border
                    elif col_letter == 'V':
                        cell_bottom.border = self.separator_border
                    else:
                        cell_bottom.border = self.border

            # Define subcategories for non-merged cells only
            subcategories = [
                (f'B{subcat_row}', "das"), (f'C{subcat_row}', "as"),
                (f'E{subcat_row}', "1 Eixo"), (f'F{subcat_row}', "2 Eixos"), (f'G{subcat_row}', "3 Eixos"),
                (f'I{subcat_row}', "2 Eixos"), (f'J{subcat_row}', "3 Eixos"), (f'K{subcat_row}', "4 Eixos"),
                (f'L{subcat_row}', "2 E"), (f'M{subcat_row}', "3 E"), (f'N{subcat_row}', "4 E"),
                (f'O{subcat_row}', "5 E"), (f'P{subcat_row}', "6 E"), (f'Q{subcat_row}', "7 E"),
                (f'R{subcat_row}', "8/9 E"), (f'S{subcat_row}', "10+ E"),
                (f'T{subcat_row}', "2 E"), (f'U{subcat_row}', "3 E ou +"),
                (f'W{subcat_row}', "% Cam"), (f'X{subcat_row}', "Caminhões"),
                (f'Y{subcat_row}', "% Carr"), (f'Z{subcat_row}', "Carretas"),
                (f'AA{subcat_row}', "% Ônib"), (f'AB{subcat_row}', "Ônibus"),
                (f'AC{subcat_row}', "% Pes"), (f'AD{subcat_row}', "Total")
            ]
            for cell_pos, value in subcategories:
                cell = self.sheet2[cell_pos]
                cell.value = value
                cell.font = Font(size=10)
                if cell_pos[0] in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                    cell.border = self.border
                elif cell_pos[0] == 'V':
                    cell.border = self.separator_border
                else:
                    cell.border = self.border
                cell.alignment = self.center_align

            das_inicio = datetime.strptime("00:00", "%H:%M")
            as_inicio = datetime.strptime("00:15", "%H:%M")
            das_fim = datetime.strptime("23:45", "%H:%M")
            row = start_row + 4
            while das_inicio <= das_fim:
                cell = self.sheet2[f'B{row}']
                cell.value = das_inicio.strftime("%H:%M")
                cell.border = self.border
                cell.alignment = self.center_align

                cell = self.sheet2[f'C{row}']
                cell.value = as_inicio.strftime("%H:%M")
                cell.border = self.border
                cell.alignment = self.center_align

                self.sheet2[f'D{row}'].value = 10  # Leves
                self.sheet2[f'H{row}'].value = 5   # VUC
                self.sheet2[f'V{row}'].value = 2   # Motos

                das_inicio += timedelta(minutes=15)
                as_inicio += timedelta(minutes=15)
                row += 1

            table_start_row = start_row + 4
            table_end_row = start_row + 99
            for row in range(table_start_row, table_end_row + 1):
                self.sheet2[f'X{row}'].value = f"=SUM(I{row}:K{row})"
                self.sheet2[f'Z{row}'].value = f"=SUM(L{row}:S{row})"
                self.sheet2[f'AB{row}'].value = f"=SUM(T{row}:U{row})"
                self.sheet2[f'AD{row}'].value = f"=SUM(X{row},Z{row},AB{row})"
                if row >= table_start_row:
                    self.sheet2[f'AE{row}'].value = f"=SUM(D{row}:H{row},V{row},AD{row})"
                self.sheet2[f'W{row}'].value = f"=IFERROR(X{row}/AE{row}, 0)"
                self.sheet2[f'W{row}'].number_format = '0.0%'
                self.sheet2[f'Y{row}'].value = f"=IFERROR(Z{row}/AE{row}, 0)"
                self.sheet2[f'Y{row}'].number_format = '0.0%'
                self.sheet2[f'AA{row}'].value = f"=IFERROR(AB{row}/AE{row}, 0)"
                self.sheet2[f'AA{row}'].number_format = '0.0%'
                self.sheet2[f'AC{row}'].value = f"=IFERROR(AD{row}/AE{row}, 0)"
                self.sheet2[f'AC{row}'].number_format = '0.0%'

                start_col = self.parent.column_to_number('D')
                end_col = self.parent.column_to_number('AE')
                for col in range(start_col, end_col + 1):
                    col_letter = get_column_letter(col)
                    cell = self.sheet2[f'{col_letter}{row}']
                    if col_letter == 'W':
                        cell.border = self.separator_border
                    else:
                        cell.border = self.border
                    cell.alignment = self.center_align

            footer_row = table_end_row + 1
            cell = self.sheet2[f'B{footer_row}']
            cell.value = "Total"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align
            self.sheet2[f'C{footer_row}'].border = self.border
            self.sheet2.merge_cells(f'B{footer_row}:C{footer_row}')

            vehicle_totals = {}
            vehicle_types = ['Leves', 'Carretinha 1E', 'Carretinha 2E', 'Carretinha 3E', 'VUC',
                            'Caminhões 2E', 'Caminhões 3E', 'Caminhões 4E', 'Carreta 2E', 'Carreta 3E',
                            'Carreta 4E', 'Carreta 5E', 'Carreta 6E', 'Carreta 7E', 'Carreta 8/9E', 'Carreta 10+E',
                            'Ônibus 2E', 'Ônibus 3E+', 'Motos']
            start_col = self.parent.column_to_number('D')
            end_col = self.parent.column_to_number('V')
            for col, vt in zip(range(start_col, end_col + 1), vehicle_types):
                col_letter = get_column_letter(col)
                cell = self.sheet2[f'{col_letter}{footer_row}']
                cell.value = f"=SUM({col_letter}{table_start_row}:{col_letter}{table_end_row})"
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align
                vehicle_totals[vt] = 0

            for col in ['W', 'Y', 'AA', 'AC']:
                cell = self.sheet2[f'{col}{footer_row}']
                cell.value = f"=IFERROR(SUM({col}{table_start_row}:{col}{table_end_row}), 0)"
                cell.font = self.header_font
                cell.fill = self.pesados_fill
                if col == 'W':
                    cell.border = self.separator_border
                else:
                    cell.border = self.border
                cell.alignment = self.center_align
                cell.number_format = '0.0%'

            for col in ['X', 'Z', 'AB', 'AD', 'AE']:
                cell = self.sheet2[f'{col}{footer_row}']
                if col == 'X':
                    cell.value = f"=SUM(I{footer_row}:K{footer_row})"
                elif col == 'Z':
                    cell.value = f"=SUM(L{footer_row}:S{footer_row})"
                elif col == 'AB':
                    cell.value = f"=SUM(T{footer_row}:U{footer_row})"
                elif col == 'AD':
                    cell.value = f"=SUM(X{footer_row},Z{footer_row},AB{footer_row})"
                elif col == 'AE':
                    cell.value = f"=SUM(D{footer_row}:H{footer_row},V{footer_row},AD{footer_row})"
                cell.font = self.header_font
                if col in ['X', 'Z', 'AB', 'AD']:
                    cell.fill = self.pesados_fill
                else:
                    cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            return footer_row + 5, movimento_concatenado, vehicle_totals, data.get("Data", "")

        def add_data(self, data):
            movimentos = data.get("Movimentos", [])
            start_row = 1
            for i, movimento in enumerate(movimentos):
                start_row, movement_name, vehicle_totals, date = self.create_movement_table(start_row, data, movimento, i)
                self.parent.vehicle_data.append((date, movement_name, vehicle_totals))

    class abaHr:
        def __init__(self, wb, parent):
            self.wb = wb
            self.parent = parent
            self.sheet3 = self.wb.create_sheet(title="Hr")
            self.header_font = Font(bold=True, size=11)
            self.header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            self.pesados_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.separator_border = Border(
                left=Side(style='medium'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.center_align = Alignment(horizontal='center', vertical='center')

        def create_movement_table(self, start_row, data, movement, movement_index):
            ponto = data.get("Ponto", "")
            movimento_concatenado = f"{ponto}{movement}" if ponto and movement else movement

            cell = self.sheet3[f'B{start_row}']
            cell.value = "Data:"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align

            cell = self.sheet3[f'C{start_row}']
            cell.value = data.get("Data", "")
            cell.border = self.border
            cell.alignment = self.center_align

            cell = self.sheet3[f'B{start_row + 1}']
            cell.value = "Movimento:"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align

            cell = self.sheet3[f'C{start_row + 1}']
            cell.value = movimento_concatenado
            cell.border = self.border
            cell.alignment = self.center_align

            header_row = start_row + 2
            subcat_row = start_row + 3
            table_columns = [
                (f'B{header_row}:C{header_row}', "Horas", None),
                (f'D{header_row}:D{subcat_row}', "Leves", None),
                (f'E{header_row}:G{header_row}', "Carretinha", None),
                (f'H{header_row}:H{subcat_row}', "VUC", None),
                (f'I{header_row}:K{header_row}', "Caminhões", None),
                (f'L{header_row}:S{header_row}', "Carreta", None),
                (f'T{header_row}:U{header_row}', "Ônibus", None),
                (f'V{header_row}:V{subcat_row}', "Motos", None),
                (f'W{header_row}:AD{header_row}', "Pesados", None),
                (f'AE{header_row}:AE{subcat_row}', "Veículos Totais", None)
            ]

            # Step 1: Apply borders, fill, and values to individual cells before merging
            for merge_range, header_value, _ in table_columns:
                start_cell, end_cell = merge_range.split(':')
                start_row_num, start_col = coordinate_to_tuple(start_cell)
                end_row_num, end_col = coordinate_to_tuple(end_cell)
                for row in range(start_row_num, end_row_num + 1):
                    for col in range(start_col, end_col + 1):
                        col_letter = get_column_letter(col)
                        cell = self.sheet3[f'{col_letter}{row}']
                        # Apply fill and border based on whether this cell is in the "Pesados" section
                        if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'] and row == header_row:
                            cell.fill = self.pesados_fill
                            cell.border = self.border
                        elif col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'] and row == subcat_row:
                            cell.fill = self.pesados_fill
                            cell.border = self.border
                        elif col_letter == 'V' and row in [header_row, subcat_row]:
                            cell.fill = self.pesados_fill
                            cell.border = self.separator_border
                        else:
                            if row == header_row:
                                cell.fill = self.header_fill
                            cell.border = self.border
                        # Set the header value only for the top-left cell
                        if row == start_row_num and col == start_col and header_value:
                            cell.value = header_value
                            cell.font = self.header_font
                            cell.alignment = self.center_align

            # Step 2: Merge the cells
            for merge_range, _, _ in table_columns:
                self.sheet3.merge_cells(merge_range)

            # Step 3: Reapply border and fill to the top-left cell of each merged area to ensure outer borders are visible
            for merge_range, _, _ in table_columns:
                start_cell, end_cell = merge_range.split(':')
                start_row_num, start_col = coordinate_to_tuple(start_cell)
                end_row_num, end_col = coordinate_to_tuple(end_cell)
                # Reapply to the top-left cell of the merged area
                cell = self.sheet3[start_cell]
                col_letter = start_cell[0]
                if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                    cell.fill = self.pesados_fill
                    cell.border = self.border
                elif col_letter == 'V':
                    cell.fill = self.pesados_fill
                    cell.border = self.separator_border
                else:
                    cell.fill = self.header_fill
                    cell.border = self.border

                # Ensure the outer borders of the merged area are visible by reapplying borders to the edges
                for row in range(start_row_num, end_row_num + 1):
                    # Leftmost column of the merged area
                    left_col_letter = get_column_letter(start_col)
                    cell_left = self.sheet3[f'{left_col_letter}{row}']
                    if left_col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_left.border = self.border
                    elif left_col_letter == 'V':
                        cell_left.border = self.separator_border
                    else:
                        cell_left.border = self.border

                    # Rightmost column of the merged area
                    right_col_letter = get_column_letter(end_col)
                    cell_right = self.sheet3[f'{right_col_letter}{row}']
                    if right_col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_right.border = self.border
                    else:
                        cell_right.border = self.border

                # Top and bottom rows of the merged area
                for col in range(start_col, end_col + 1):
                    col_letter = get_column_letter(col)
                    # Top row
                    cell_top = self.sheet3[f'{col_letter}{start_row_num}']
                    if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_top.border = self.border
                    elif col_letter == 'V':
                        cell_top.border = self.separator_border
                    else:
                        cell_top.border = self.border

                    # Bottom row
                    cell_bottom = self.sheet3[f'{col_letter}{end_row_num}']
                    if col_letter in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                        cell_bottom.border = self.border
                    elif col_letter == 'V':
                        cell_bottom.border = self.separator_border
                    else:
                        cell_bottom.border = self.border

            # Define subcategories for non-merged cells only
            subcategories = [
                (f'B{subcat_row}', "das"), (f'C{subcat_row}', "as"),
                (f'E{subcat_row}', "1 Eixo"), (f'F{subcat_row}', "2 Eixos"), (f'G{subcat_row}', "3 Eixos"),
                (f'I{subcat_row}', "2 Eixos"), (f'J{subcat_row}', "3 Eixos"), (f'K{subcat_row}', "4 Eixos"),
                (f'L{subcat_row}', "2 E"), (f'M{subcat_row}', "3 E"), (f'N{subcat_row}', "4 E"),
                (f'O{subcat_row}', "5 E"), (f'P{subcat_row}', "6 E"), (f'Q{subcat_row}', "7 E"),
                (f'R{subcat_row}', "8/9 E"), (f'S{subcat_row}', "10+ E"),
                (f'T{subcat_row}', "2 E"), (f'U{subcat_row}', "3 E ou +"),
                (f'W{subcat_row}', "% Cam"), (f'X{subcat_row}', "Caminhões"),
                (f'Y{subcat_row}', "% Carr"), (f'Z{subcat_row}', "Carretas"),
                (f'AA{subcat_row}', "% Ônib"), (f'AB{subcat_row}', "Ônibus"),
                (f'AC{subcat_row}', "% Pes"), (f'AD{subcat_row}', "Total")
            ]
            for cell_pos, value in subcategories:
                cell = self.sheet3[cell_pos]
                cell.value = value
                cell.font = Font(size=10)
                if cell_pos[0] in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                    cell.border = self.border
                elif cell_pos[0] == 'V':
                    cell.border = self.separator_border
                else:
                    cell.border = self.border
                cell.alignment = self.center_align

            das_inicio = datetime.strptime("00:00", "%H:%M")
            as_inicio = datetime.strptime("01:00", "%H:%M")
            das_fim = datetime.strptime("23:00", "%H:%M")
            row = start_row + 4
            while das_inicio <= das_fim:
                cell = self.sheet3[f'B{row}']
                cell.value = das_inicio.strftime("%H:%M")
                cell.border = self.border
                cell.alignment = self.center_align

                cell = self.sheet3[f'C{row}']
                cell.value = as_inicio.strftime("%H:%M")
                cell.border = self.border
                cell.alignment = self.center_align

                das_inicio += timedelta(hours=1)
                as_inicio += timedelta(hours=1)
                row += 1

            try:
                relatorio_sheet = self.wb['Relatório']
                table_start_row = start_row + 4
                table_end_row = start_row + 27
                for hr_row, hour in enumerate(range(24), table_start_row):
                    rel_row_start = 5 + (hour * 4)
                    col_mapping = {
                        'D': 'D', 'H': 'H', 'V': 'V',
                        'E': 'E', 'F': 'F', 'G': 'G',
                        'I': 'I', 'J': 'J', 'K': 'K',
                        'L': 'L', 'M': 'M', 'N': 'N', 'O': 'O', 'P': 'P', 'Q': 'Q', 'R': 'R', 'S': 'S',
                        'T': 'T', 'U': 'U',
                        'X': 'X', 'Z': 'Z', 'AB': 'AB', 'AD': 'AD', 'AE': 'AE'
                    }
                    for col in col_mapping:
                        self.sheet3[f'{col}{hr_row}'].value = (
                            f"=SUM('Relatório'!{col}{rel_row_start}:'Relatório'!{col}{rel_row_start+3})"
                        )
            except KeyError:
                print("Warning: 'Relatório' sheet not found in workbook.")
                cell = self.sheet3[f'B{table_start_row}']
                cell.value = "Folha 'Relatório' não encontrada"
                cell.border = self.border
                cell.alignment = self.center_align

            for row in range(table_start_row, table_end_row + 1):
                self.sheet3[f'X{row}'].value = f"=SUM(I{row}:K{row})"
                self.sheet3[f'Z{row}'].value = f"=SUM(L{row}:S{row})"
                self.sheet3[f'AB{row}'].value = f"=SUM(T{row}:U{row})"
                self.sheet3[f'AD{row}'].value = f"=SUM(X{row},Z{row},AB{row})"
                self.sheet3[f'AE{row}'].value = f"=SUM(D{row}:H{row},V{row},AD{row})"
                self.sheet3[f'W{row}'].value = f"=IFERROR(X{row}/AE{row}, 0)"
                self.sheet3[f'W{row}'].number_format = '0.0%'
                self.sheet3[f'Y{row}'].value = f"=IFERROR(Z{row}/AE{row}, 0)"
                self.sheet3[f'Y{row}'].number_format = '0.0%'
                self.sheet3[f'AA{row}'].value = f"=IFERROR(AB{row}/AE{row}, 0)"
                self.sheet3[f'AA{row}'].number_format = '0.0%'
                self.sheet3[f'AC{row}'].value = f"=IFERROR(AD{row}/AE{row}, 0)"
                self.sheet3[f'AC{row}'].number_format = '0.0%'

                start_col = self.parent.column_to_number('D')
                end_col = self.parent.column_to_number('AE')
                for col in range(start_col, end_col + 1):
                    col_letter = get_column_letter(col)
                    cell = self.sheet3[f'{col_letter}{row}']
                    if col_letter == 'W':
                        cell.border = self.separator_border
                    else:
                        cell.border = self.border
                    cell.alignment = self.center_align

            footer_row = table_end_row + 1
            cell = self.sheet3[f'B{footer_row}']
            cell.value = "Total"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align
            self.sheet3[f'C{footer_row}'].border = self.border
            self.sheet3.merge_cells(f'B{footer_row}:C{footer_row}')

            start_col = self.parent.column_to_number('D')
            end_col = self.parent.column_to_number('V')
            for col in range(start_col, end_col + 1):
                col_letter = get_column_letter(col)
                cell = self.sheet3[f'{col_letter}{footer_row}']
                cell.value = f"=SUM({col_letter}{table_start_row}:{col_letter}{table_end_row})"
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            for col in ['W', 'Y', 'AA', 'AC']:
                cell = self.sheet3[f'{col}{footer_row}']
                cell.value = f"=IFERROR(SUM({col}{table_start_row}:{col}{table_end_row}), 0)"
                cell.font = self.header_font
                cell.fill = self.pesados_fill
                if col == 'W':
                    cell.border = self.separator_border
                else:
                    cell.border = self.border
                cell.alignment = self.center_align
                cell.number_format = '0.0%'

            for col in ['X', 'Z', 'AB', 'AD', 'AE']:
                cell = self.sheet3[f'{col}{footer_row}']
                if col == 'X':
                    cell.value = f"=SUM(I{footer_row}:K{footer_row})"
                elif col == 'Z':
                    cell.value = f"=SUM(L{footer_row}:S{footer_row})"
                elif col == 'AB':
                    cell.value = f"=SUM(T{footer_row}:U{footer_row})"
                elif col == 'AD':
                    cell.value = f"=SUM(X{footer_row},Z{footer_row},AB{footer_row})"
                elif col == 'AE':
                    cell.value = f"=SUM(D{footer_row}:H{footer_row},V{footer_row},AD{footer_row})"
                cell.font = self.header_font
                if col in ['X', 'Z', 'AB', 'AD']:
                    cell.fill = self.pesados_fill
                else:
                    cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align

            return footer_row + 5

        def add_data(self, data):
            movimentos = data.get("Movimentos", [])
            start_row = 1
            for i, movimento in enumerate(movimentos):
                start_row = self.create_movement_table(start_row, data, movimento, i)

    def add_data(self, data):
        self.data = data
        self.entrada.add_data(data)
        self.relatorio.add_data(data)
        self.hr.add_data(data)

        duration_days_str = data.get("Duração em dias", 1)
        try:
            duration_days = int(duration_days_str)
        except (ValueError, TypeError):
            duration_days = 1
            print(f"Warning: Invalid 'Duração em dias' value '{duration_days_str}', defaulting to 1.")

        days = []
        try:
            initial_date = datetime.strptime(data.get("Data", ""), "%d-%m-%Y")
        except ValueError:
            try:
                initial_date = datetime.strptime(data.get("Data", ""), "%Y-%m-%d")
            except ValueError:
                initial_date = datetime.now()
                print(f"Warning: Invalid date format, using current date.")
        days.append(initial_date.strftime("%Y-%m-%d"))

        if duration_days > 1:
            for day in range(1, duration_days):
                relatorio_copy = self.abaRelatorio(self.wb, self)
                relatorio_copy.sheet2.title = f"Relatório ({day})"
                hr_copy = self.abaHr(self.wb, self)
                hr_copy.sheet3.title = f"Hr ({day})"
                copy_data = data.copy()
                new_date = initial_date + timedelta(days=day)
                copy_data["Data"] = new_date.strftime("%Y-%m-%d")
                days.append(new_date.strftime("%Y-%m-%d"))
                relatorio_copy.add_data(copy_data)
                hr_copy.add_data(copy_data)

        self.resumo.add_data(self.vehicle_data, days)

    def save(self):
        for sheet in self.wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min((max_length + 2), 100)
                sheet.column_dimensions[column].width = adjusted_width
        self.wb.save(f"output/{self.filename}")
        print(f"Planilha salva como {self.filename}")